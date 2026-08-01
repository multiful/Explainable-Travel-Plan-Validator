#!/usr/bin/env python3
"""제주특별자치도_통합.xlsx의 위도/경도 결측 행을 Kakao API로 지오코딩해 채운다.

전략: 1차 = 도로명주소 지오코딩(KakaoLocalClient.geocode_address),
      2차(1차 실패 시) = "상호명 도로명주소" 키워드 검색(search_keyword) 폴백.
원본 파일은 .bak으로 백업한 뒤 같은 경로에 덮어쓴다 (git 미추적 파일이라 백업 필수).
"""
from __future__ import annotations

import shutil
import sys
import time
from pathlib import Path

import openpyxl
from dotenv import load_dotenv

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))
load_dotenv(ROOT / ".env")

from src.data.kakao_local import KakaoLocalClient  # noqa: E402

XLSX_PATH = ROOT / "제주특별자치도_통합.xlsx"
SHEET_NAME = "제주특별자치도_통합"


def main() -> None:
    if not XLSX_PATH.exists():
        raise SystemExit(f"파일 없음: {XLSX_PATH}")

    client = KakaoLocalClient.from_env()
    if not client.enabled:
        raise SystemExit("KAKAO_REST_API_KEY 미설정 — .env 확인 필요")

    backup_path = XLSX_PATH.with_suffix(".xlsx.bak")
    shutil.copy2(XLSX_PATH, backup_path)
    print(f"백업 생성: {backup_path}")

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb[SHEET_NAME]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {name: idx + 1 for idx, name in enumerate(headers)}  # 1-based

    recovered_by_address = 0
    recovered_by_keyword = 0
    still_missing: list[tuple[int, str, str]] = []

    rows = list(ws.iter_rows(min_row=2))
    total_missing = sum(1 for r in rows if r[col["위도"] - 1].value is None)
    print(f"결측 좌표 행: {total_missing}건")

    for i, row in enumerate(rows, start=1):
        lat_cell = row[col["위도"] - 1]
        lng_cell = row[col["경도"] - 1]
        if lat_cell.value is not None:
            continue

        name = row[col["상호명"] - 1].value or ""
        address = row[col["도로명주소"] - 1].value or ""
        row_id = row[col["ID"] - 1].value

        coords = client.geocode_address(address)
        source = "address"
        if coords is None:
            place = client.search_keyword(f"{name} {address}") or client.search_keyword(name)
            coords = (place.lat, place.lng) if place else None
            source = "keyword"

        if coords is not None:
            lat, lng = coords
            lat_cell.value = lat
            lng_cell.value = lng
            if source == "address":
                recovered_by_address += 1
            else:
                recovered_by_keyword += 1
        else:
            still_missing.append((row_id, name, address))

        if i % 50 == 0:
            print(f"  진행: {i}/{len(rows)}")
        time.sleep(0.05)  # Kakao 초당 호출 제한 여유

    wb.save(XLSX_PATH)

    recovered = recovered_by_address + recovered_by_keyword
    print("\n=== 결과 ===")
    print(f"주소 지오코딩 성공: {recovered_by_address}")
    print(f"키워드 검색 폴백 성공: {recovered_by_keyword}")
    print(f"총 복구: {recovered} / {total_missing}")
    print(f"여전히 결측: {len(still_missing)}")
    for row_id, name, address in still_missing:
        print(f"  - ID={row_id} {name} | {address}")


if __name__ == "__main__":
    main()
